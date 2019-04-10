import json
import os
import glob
from json import JSONEncoder
import pandas as pd
from math import isnan

# from pandas.io.json import json_normalize


job_definition_file = f'../../baselines.json'
job_definition = json.loads(open(job_definition_file, 'r').read())

report_folder = f'/home/setepenre/mlperf_output/'
report_folder = f'/home/user1/mlperf/results'


class TimeSeries:
    def __init__(self, s=0, count=0):
        self.sum = s
        self.count = count
        self.values = []

    @property
    def avg(self):
        try:
            return self.sum / self.count
        except ZeroDivisionError:
            return float('inf')

    def max(self):
        try:
            return max(self.values)
        except ValueError:
            return float('NaN')

    def __iadd__(self, other):
        return self.update(other)

    def to_dict(self):
        return {'avg': self.avg, 'count': self.count}

    def update(self, other):
        self.sum += other
        self.count += 1
        self.values.append(other)
        return self


class AverageEncoder(JSONEncoder):
    def default(self, o):
        return o.to_dict()


def score(report):
    values = report['train_item']
    avg, range = values['avg'], values['range']
    # min, max = values['min'], values['max']

    return avg / range


def report():
    filename = os.environ.get('REPORT_PATH', '/home/user1/mlperf/output/bench_result.csv')

    report_file = open(filename, 'r')
    reports = report_file.read()[:-1] + ']'
    report_file.close()

    for report in reports:
        print(json.dumps(report, indent=4))


def read_results(report_name):
    return json.loads(open(report_name, 'r').read()[:-1] + ']')


def isnot_nan(x):
    return not isnan(x)

# d = pd.DataFrame({
#     'vendor1': {
#         'conv': [1, 2, 3],
#         'mnist': [1, 2, 3]
#     },
#     'vendor2': { # bench_name: [device_score...]
#         'conv': [1, 2],
#         'mnist': [1, 2]
#     }
# })
# Get overall result
# print(d.applymap(lambda x: sum(x)))


results_break_down = {}
overall_results = {}
vendor_to_device_count = {}

for vendor_name in os.listdir(report_folder):
    print(f'Processing vendor: {vendor_name}')
    baselines_results = glob.glob(f'{report_folder}/{vendor_name}/baselines_*')

    device_count = len(baselines_results)

    vendor_score = 0
    gpu_count = 0

    # Vendor -> Bench Name -> Score
    breaked_down = {}
    results_break_down[vendor_name] = breaked_down
    vendor_to_device_count[vendor_name] = 0

    for idx, device_reports in enumerate(baselines_results):
        # print(f'{">" * 2} Processing Report {device_reports}')
        reports = read_results(device_reports)
        vendor_to_device_count[vendor_name] += 1
        bench_name = None

        device_score = 0
        bench_count = 0

        for bench_result in reports:
            bench_name = bench_result['name']
            if 'fp16' in bench_result:
                bench_name += '_fp16'
            elif 'opt_level' in bench_result:
                if bench_result['opt_level'] != 'O0':
                    bench_name += '_fp16'

            if bench_name not in breaked_down:
                breaked_down[bench_name] = [TimeSeries()]

            if len(breaked_down[bench_name]) == idx:
                breaked_down[bench_name].append(TimeSeries())

            bench_score = score(bench_result)
            breaked_down[bench_name][idx].update(bench_score)

        gpu_count += 1
        print()


print(json.dumps(results_break_down, cls=AverageEncoder, indent=4))

original = pd.DataFrame(results_break_down)
per_device = original.applymap(lambda x: [i.max() for i in x])

cols = 'ABCDEFGHIJKLMNOPQRSTUVWXYZ'


from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import colors
from openpyxl.styles import Font, Color


wb = Workbook()
ws = wb.active

print('Per Device Performance')
shape = per_device.shape
print(per_device)


ws['A1'] = 'Per Device Performance'
data = dataframe_to_rows(per_device)
for r, row in enumerate(data, 2):
    offset = 0

    for c, value in enumerate(row, 1):
        device_count = 1
        if c != 0:
            vendor = per_device.columns[c - 2]
            device_count = vendor_to_device_count[vendor]

        if isinstance(value, list):
            for i, v in enumerate(value):

                if isnan(v):
                    v = 'nan'

                ws.cell(row=r, column=c + i + offset, value=v)
        else:
            ws.cell(row=r, column=c + offset, value=value)

        offset += device_count - 1


cell = f'A{shape[0] + 5}'
ws[cell] = 'Overall Per Bench Performance'

print()
overall = per_device.applymap(lambda x: sum(filter(isnot_nan, x)) / len(list(filter(isnot_nan, x))))

print('Overall Per Bench Performance')
print(overall)

s = shape[0] + 5
data = dataframe_to_rows(overall)
for r, row in enumerate(data, 2):
    for c, value in enumerate(row, 1):
        ws.cell(row=r + s, column=c, value=value)


wb.save('report.xlsx')




#print(json.dumps(results_break_down, cls=AverageEncoder, indent=4))
#print(json.dumps(overall_results, indent=4))



# add_timestep	algo	all.avg	all.count	all.max	all.min	all.sd	all.unit
# alpha	arch	batch_loss	batch_size	beam_size	beta1	bptt	bucketing
# checkpoint	checkpoint_interval	checkpoint_model_dir	clip	clip_param
# content_weight	cov_penalty_factor	cpu_cores	cuda	cuda_deterministic
# cudnn	data	dataroot	dataset	dataset_dir	devices	disable_eval
# dist_url	dropout	dtype	dynamic_loss_scale	emsize	entropy_coef
# env_name	epoch_loss	epochs	eps
# eval.avg	eval.count	eval.max	eval.min	eval.sd	eval.unit
# eval_batch_size	eval_interval	evaluation	factors	fp16
# gamma	gpu	grad_clip	hostname	imageSize	image_size	index
# iteration	jr_id	keep_checkpoints	layers	learning_rate	len_norm_const
# len_norm_factor
# loading_data.avg	loading_data.count	loading_data.max	loading_data.min	loading_data.sd	loading_data.unit
# local_rank	log_dir	log_interval	lr	math	max_grad_norm	max_length_train_item	max_length_val
# max_size
# metrics.errD_real	metrics.errG	metrics.errG_fake	metrics.psnr	metrics.val_loss	metrics.value_loss
# min_length_train_item	min_length_val	model	model_config	name	ndf	negative_samples	netD	netG	ngf
# ngpu	nhid	nlayers	no_checks	no_save	num_env_steps	num_mini_batch	num_processes	num_steps	number
# nz	onnx_export	opt_level	optimization_config	outf	port	ppo_epoch	print_freq	processes	prof
# rank	recurrent_policy	render	repeat	results_dir	resume	save	save_all	save_dir	save_freq
# save_interval	save_model_dir	smoothing	start_epoch	static_loss_scale	style_image	style_size	style_weight
# subcommand	target_bleu	tau	testBatchSize	threshold	tied	topk
# train_item.avg	train_item.count	train_item.max	train_item.min	train_item.sd	train_item.unit
# train_item_item.avg	train_item_item.max	train_item_item.min	train_item_item.range	train_item_item.unit
# unique_id	upscale_factor	use_gae	use_linear_clip_decay	use_linear_lr_decay	value_loss_coef
# vcd	version	vis	vis_interval	workers	world_size


#df = json_normalize(reports)
#df.to_csv('all_aggregated.csv')
selected_columns = {
    'gpu', 'hostname', 'index', 'name'
    ,'train_item.avg'
    ,'train_item.count'
    ,'train_item.max'
    ,'train_item.min'
    ,'train_item.sd'
    ,'train_item.unit'
    ,'train_item_item.avg'
    ,'train_item_item.max'
    ,'train_item_item.min'
    ,'train_item_item.range'
    ,'train_item_item.unit'
    ,'unique_id'
    ,'version'
    ,'workers'
    ,'world_size'
    ,'vcd'
    ,'devices'
    ,'cpu_cores'
    ,'batch_size'
    ,'all.avg'
    ,'all.count'
    ,'all.max'
    ,'all.min'
    ,'all.sd'
    ,'eval.avg'
    ,'eval.count'
    ,'eval.max'
    ,'eval.min'
    ,'eval.sd'
    ,'fp16'
    ,'loading_data.avg'
    ,'loading_data.count'
    ,'loading_data.max'
    ,'loading_data.min'
    ,'loading_data.sd'
}

#perf_report = df[selected_columns]
#perf_report.to_csv('perf_report.csv', index=False)


